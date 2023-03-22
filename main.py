#biblioteca para excel
from openpyxl import load_workbook
#Bibliotecas para obtener fecha del sistema
import datetime
from datetime import datetime
import zipfile
import os
from os import walk

#Biblioteca del sistema operativo
from clases.Nomina12 import * 
from clases.Comprobante import * 
from clases.Percepciones import * 
from clases.Percepcion import * 
from clases.Deducciones import * 
from clases.Deduccion import * 
from clases.OtrosPagos import * 
from clases.Incapacidades import * 
from clases.Empleado import * 

def buscar(objBuscar,idNomina):
    for obj in objBuscar:
        if obj.getIDNOMINA() == idNomina:
            return obj.getID()

def buscarEmpleado(lista, RFC, bandera):
    if bandera:
        for objeto in lista:
            if objeto.getRFC() == RFC:
                return objeto.getNEMPLEADO();
    else:
        for objeto in lista:
            if objeto.getNEMPLEADO() == RFC:
                return objeto.getRFC();

def crearArchivos(path, file):
    
    wb = load_workbook(file)
    #HOJA INFORMACION GENERAL COMPROBANTE
    sheet_general = wb ['DC']
    #HOJA COMPLEMENTO NÓMINA
    sheet_complemento = wb['CNR']
    #HOJA INFORMACION DEL EMISOR
    sheet_emisor = wb['EM']
    #HOJA COMPLEMENTO NÓMINA EMISOR
    sheet_emisor_complemento = wb['CNE']
    #HOJA CONCEPTOS DEL COMPRONBANTE
    sheet_conceptos = wb['CN']
    #HOJA LISTA DE CORREOS DESTINATARIOS
    sheet_correo = wb['MI']
    #HOJA INFORMACION DEL CFDI RELACIONADO
    sheet_cfdi = wb['CRL']
    #HOJA INFORMACION DEL RECEPTOR
    sheet_receptor = wb['RC']
    #HOJA INFORMACION CAMPOS OPCIONALES 
    sheet_opcionales = wb['OP']
    #HOJA PERCEPCIONES
    sheet_percepciones = wb['CNP']
    #HOJA DETALLE DE PERCEPCIONES
    sheet_percepcion = wb['NPD']
    #HOJA INFORMACION DE DEDUCCIONES
    sheet_deduccion = wb['CND']
    #HOJA DETALLE DE DEDUCCIONES
    sheet_deducciones = wb['NDD']
    #HOJA INFORMACION DE OTROS PAGOS
    sheet_pagos = wb['NOP']
    #HOJA INFORMACION DE INCAPACIDADES
    sheet_incapacidad = wb['CNI']    
    
    listEmpleados = []
    
    count =  7
    
    if sheet_complemento.cell(row = count, column = 1).value != None:
        while sheet_complemento.cell(row = count, column = 2).value != None:
            obj = Empleado(sheet_receptor.cell(row = count, column = 1).value ,sheet_complemento.cell(row = count, column = 18).value)
            listEmpleados.append(obj)
            count = count + 1
            
            
    
    
    #----------------------------------------COMPROBANTES DC -------------------------------------------
    fileaux = path + "/M4T_COMPROBANTES_DC_33.csv"
    count =  7
    idcomprobante = 1
    rComprobantes = []
    if sheet_general.cell(row = count, column = 1).value != None:
        while sheet_general.cell(row = count, column = 1).value != None: 
            obj = Comprobante()
            obj.setIDCOMPROBANTE(idcomprobante)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_complemento.cell(row = count, column = 1).value,True))
            obj.setSERIE(sheet_general.cell(row = count, column = 1).value)
            obj.setFOLIO(sheet_general.cell(row = count, column = 2).value)
            obj.setFECHA(sheet_general.cell(row = count, column = 3).value)
            obj.setFORMADEPAGO(sheet_general.cell(row = count, column = 4).value)
            obj.setSUBTOTAL(sheet_general.cell(row = count, column = 5).value)
            obj.setDESCUENTO(sheet_general.cell(row = count, column = 6).value)
            obj.setTOTAL(sheet_general.cell(row = count, column = 9).value)
            obj.setMETODODEPAGO(sheet_general.cell(row = count, column = 11).value)
            obj.setTIPODECOMPROBANTE(sheet_general.cell(row = count, column = 10).value)
            obj.setMONEDA(sheet_general.cell(row = count, column = 7).value)
            obj.setLUGAREXPEDICION(sheet_general.cell(row = count, column = 12).value)
            obj.setRFCEMISOR(sheet_emisor.cell(row = count, column = 1).value)
            obj.setNOMBREEMISOR(sheet_emisor.cell(row = count, column = 2).value)
            obj.setREGIMENFISCALEMISOR(sheet_emisor_complemento.cell(row = count, column = 2).value)
            obj.setREGISTROPATRONAL(sheet_emisor_complemento.cell(row = count, column = 4).value)
            obj.setORIGENRECURSO(sheet_emisor_complemento.cell(row = count, column = 6).value)
            obj.setCANTIDAD(sheet_conceptos.cell(row = count, column = 3).value)
            obj.setDESCRIPCION(sheet_conceptos.cell(row = count, column = 5).value)
            obj.setVALORUNITARIO(sheet_conceptos.cell(row = count, column = 6).value)
            obj.setIMPORTE(sheet_conceptos.cell(row = count, column = 7).value)
            obj.setPARA(sheet_correo.cell(row = count, column = 2).value)
            obj.setCC(sheet_correo.cell(row = count, column = 3).value)
            obj.setCCO(sheet_correo.cell(row = count, column = 4).value)
            obj.setMONTORECURSOPROPIO(sheet_emisor_complemento.cell(row = count, column = 7).value)
            obj.setTIPORELACION(sheet_general.cell(row = count, column = 16).value)
            obj.setUUID(sheet_general.cell(row = count, column = 17).value)
            obj.setUSODECFDI(sheet_receptor.cell(row = count, column = 3).value)
            obj.setCLAVEPRODSERV(sheet_conceptos.cell(row = count, column = 2).value)
            obj.setCLAVEUNIDAD(sheet_conceptos.cell(row = count, column = 4).value)
            
            obj.setEXPORTACION(sheet_general.cell(row= count, column=18).value)

            obj.setobj_objetos(sheet_general.cell(row=count, column=19).value)
            
            
            idcomprobante = idcomprobante + 1
            count = count + 1
            rComprobantes.append(obj)
            
        dc = open(fileaux,"w")
        cad = "IDCOMPROBANTE,IDNOMINA,VERSION,SERIE,FOLIO,FECHA,FORMADEPAGO,SUBTOTAL,DESCUENTO,TOTAL,METODODEPAGO,TIPODECOMPROBANTE,MONEDA,LUGAREXPEDICION,RFCEMISOR,NOMBREEMISOR,REGIMENFISCALEMISOR,CURPEMISOR,REGISTROPATRONAL,RFCPATRONORIGEN,ORIGENRECURSO,CANTIDAD,DESCRIPCION,VALORUNITARIO,IMPORTE,PARA,CC,CCO,PROCESADO,ESTATUS,ID_UNIDAD_ADM,IDUSUARIO,FG,ST,MONTORECURSOPROPIO,IDSUCURSAL,CONDICIONESDEPAGO,TIPORELACION,UUID,USODECFDI,CLAVEPRODSERV,NOIDENTIFICACION,CLAVEUNIDAD,BASE,IMPUESTO,TIPOFACTOR,TASAOCUOTA,UUIDSUCURSAL,EXPORTACION,OBJ_\n"
        dc.write(cad)
        for lista in rComprobantes:
            dc.write(lista.getinfo(1))
        dc.close()
        #----------------------------------------Fin Comprobante ----------------------------------------
    #----------------------------------------Nomina 12 -------------------------------------------
    fileaux = path + "/M4T_NOMINA12_CNR_33.csv"
    
    count =  7
    idcomprobante = 1
    nomina = []
    
    if sheet_complemento.cell(row = count, column = 1).value != None:
        while sheet_complemento.cell(row = count, column = 2).value != None:
            obj = Nomina12()
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_complemento.cell(row = count, column = 1).value,True))
            obj.setIDCOMPROBANTE(idcomprobante)
            obj.setRFCRECEPTOR(sheet_receptor.cell(row = count, column = 1).value)
            obj.setNOMBRERECEPTOR(sheet_receptor.cell(row = count, column = 2).value)
            obj.setTIPONOMINA(sheet_complemento.cell(row = count, column = 2).value)
            obj.setFECHAPAGO(sheet_complemento.cell(row = count, column = 3).value)
            obj.setFECHAINICIALPAGO(sheet_complemento.cell(row = count, column = 4).value)
            obj.setFECHAFINALPAGO(sheet_complemento.cell(row = count, column = 5).value)
            obj.setNUMDIASPAGADOS(sheet_complemento.cell(row = count, column = 6).value)
            obj.setTOTALPERCEPCIONES(sheet_complemento.cell(row = count, column = 7).value)
            obj.setTOTALDEDUCCIONES(sheet_complemento.cell(row = count, column = 8).value)
            obj.setTOTALOTROSPAGOS(sheet_complemento.cell(row = count, column = 9).value)
            obj.setCURP(sheet_complemento.cell(row = count, column = 10).value)
            obj.setNUMSEGURIDADSOCIAL(sheet_complemento.cell(row = count, column = 11).value)
            obj.setANTIGUEDAD(sheet_complemento.cell(row = count, column = 13).value)
            obj.setTIPOCONTRATO(sheet_complemento.cell(row = count, column = 14).value)
            obj.setSINDICALIZADO(sheet_complemento.cell(row = count, column = 15).value)
            obj.setTIPOJORNADA(sheet_complemento.cell(row = count, column = 16).value)
            obj.setTIPOREGIMEN(sheet_complemento.cell(row = count, column = 17).value)
            obj.setNUMEMPLEADO(sheet_complemento.cell(row = count, column = 18).value)
            obj.setDEPARTAMENTO(sheet_complemento.cell(row = count, column = 19).value)
            obj.setPUESTO(sheet_complemento.cell(row = count, column = 20).value)
            obj.setRIESGOPUESTO(sheet_complemento.cell(row = count, column = 21).value)
            obj.setPERIODICIDADPAGO(sheet_complemento.cell(row = count, column = 22).value)
            obj.setBANCO(sheet_complemento.cell(row = count, column = 23).value)
            obj.setCUENTABANCARIA(sheet_complemento.cell(row = count, column = 24).value)
            obj.setSALARIOBASECOTAPOR(sheet_complemento.cell(row = count, column = 25).value)
            obj.setSALARIODIARIOINTEGRADO(sheet_complemento.cell(row = count, column = 26).value)
            obj.setCLAVEENTFED(sheet_complemento.cell(row = count, column = 27).value)
            obj.setN_UNIDAD_ADM(sheet_opcionales.cell(row = count, column = 4).value)
            obj.setCOD_PUESTO_CVE_ACT(sheet_opcionales.cell(row = count, column = 11).value)
            obj.setN_PUESTO_ACT_ASOC_PROG(sheet_opcionales.cell(row = count, column = 13).value)
            obj.setTIPO_CONTRATACION_SUBPROG(sheet_opcionales.cell(row = count, column = 16).value)
            obj.setPERIODO_PAGO(sheet_opcionales.cell(row = count, column = 18).value)
            obj.setFECHAINICIOREALLABORAL(sheet_complemento.cell(row = count, column = 12).value)
            obj.setDEDUCCIONESAD(sheet_complemento.cell(row = count, column = 28).value)
            obj.setSERIE(sheet_complemento.cell(row = count, column = 29).value)
            nomina.append(obj)
            idcomprobante = idcomprobante + 1
            count = count + 1
            
        dc = open(fileaux,"w")
        cad = "IDNOMINA,IDCOMPROBANTE,RFCRECEPTOR,NOMBRERECEPTOR,TIPONOMINA,FECHAPAGO,FECHAINICIALPAGO,FECHAFINALPAGO,NUMDIASPAGADOS,TOTALPERCEPCIONES,TOTALDEDUCCIONES,TOTALOTROSPAGOS ,CURP,NUMSEGURIDADSOCIAL,ANTIGUEDAD,TIPOCONTRATO,SINDICALIZADO,TIPOJORNADA,TIPOREGIMEN,NUMEMPLEADO,DEPARTAMENTO,PUESTO,RIESGOPUESTO,PERIODICIDADPAGO,BANCO,CUENTABANCARIA,SALARIOBASECOTAPOR,SALARIODIARIOINTEGRADO,CLAVEENTFED,IDUSUARIO,FEC_ULT_ACTUALIZACION,COMENT,SERIE,SECT_PRES,N_SECT_PRES,UNIDAD_ADM,N_UNIDAD_ADM,ZONA_PAGADORA,NUM_PLAZA,UNIVERSO,NIVEL_SALARIAL,COD_PUESTO_CVE_ACT,GRADO,N_PUESTO_ACT_ASOC_PROG,SINDICATO,COMISION_SIND,TIPO_CONTRATACION_SUBPROG,PERIODO_PAGO,LEY_CUMPLEANIOS,AVISOS,NUM_CUENTA,INICIO_DE_LAB,TURNO_LAB,CODIGO_CONASA,N_CODIGO_CONASA,FG,ST,PERIODO_CONTRATACION,TIPO_NOMINA,FECHAINICIOREALLABORAL,DEDUCCIONESAD,CEDULAPROF,FOLIO_CFDI,VERSION,FECHATIMBRADO,SELLOCFD,NOCERTIFICADOSAT,SELLOSAT\n"
        dc.write(cad)
        for lista in nomina:
            dc.write(lista.getinfo(1))
        dc.close()
      #-------------------------------------------FIN DE NOMINA 12------------------------------------------#
    
    #----------------------------------------INICIO DE PERCEPCIONES -------------------------------------------
    fileaux = path + "/M4T_PERCEPCIONES_CNP_33.csv"
    count =  7
    Listpercepciones = []
    idPercepciones = 1
    
    if sheet_percepciones.cell(row = count, column = 1).value != None:
        while sheet_percepciones.cell(row = count, column = 1).value != None:
            obj = Percepciones()
            obj.setIDPERCEPCIONES(idPercepciones)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_percepciones.cell(row = count, column = 1).value,True))
            obj.setTOTALSUELDOS(sheet_percepciones.cell(row = count, column = 2).value)
            obj.setTOTALSEPARACIONINDEM(sheet_percepciones.cell(row = count, column = 3).value)
            obj.setTOTALJUBILACIONPENSIONRE(sheet_percepciones.cell(row = count, column = 4).value)
            obj.setTOTALGRAVADO(sheet_percepciones.cell(row = count, column = 5).value)
            obj.setTOTALEXENTO(sheet_percepciones.cell(row = count, column = 6).value)
            obj.setTOTALUNAEXHIBICION(sheet_percepciones.cell(row = count, column = 7).value)
            obj.setINGRESOACUMULABLEPEN(sheet_percepciones.cell(row = count, column = 10).value)
            obj.setINGRESONOACUMULABLEPEN(sheet_percepciones.cell(row = count, column = 11).value)
            obj.setTOTALPAGADO(sheet_percepciones.cell(row = count, column = 12).value)
            obj.setNUMANIOSSERVICIO(sheet_percepciones.cell(row = count, column = 13).value)
            obj.setULTIMOSUELDOMENSORD(sheet_percepciones.cell(row = count, column = 14).value)
            obj.setINGRESOACUMULABLEINDEM(sheet_percepciones.cell(row = count, column = 15).value)
            obj.setINGRESONOACUMULABLEINDEM(sheet_percepciones.cell(row = count, column = 16).value)
            obj.setTOTALPARCIALIDAD(sheet_percepciones.cell(row = count, column = 8).value)
            obj.setMONTODIARIO(sheet_percepciones.cell(row = count, column = 9).value)
            Listpercepciones.append(obj)
            idPercepciones = idPercepciones + 1
            count = count + 1
        
        dc = open(fileaux,"w")
        cad = "IDPERCEPCIONES,IDNOMINA,TOTALSUELDOS,TOTALSEPARACIONINDEM,TOTALJUBILACIONPENSIONRE,TOTALGRAVADO,TOTALEXENTO,TOTALUNAEXHIBICION,INGRESOACUMULABLEPEN,INGRESONOACUMULABLEPEN,TOTALPAGADO,NUMANIOSSERVICIO,ULTIMOSUELDOMENSORD,INGRESOACUMULABLEINDEM,INGRESONOACUMULABLEINDEM,IDUSUARIO,FG,ST,TOTALPARCIALIDAD,MONTODIARIO\n"
        dc.write(cad)
        for lista in Listpercepciones:
            dc.write(lista.getinfo(1))
        dc.close()
    #----------------------------------------INICIO DE PERCEPCION -------------------------------------------
    fileaux = path + "/M4T_PERCEPCION_NPD_33.csv"
    count =  7
    Listpercepcion = []
    idPercepcion = 1
        
    if sheet_percepcion.cell(row = count, column = 1).value != None:
        while sheet_percepcion.cell(row = count, column = 1).value != None:
            obj = Percepcion()
            obj.setIDPERCEPCION(sheet_percepcion.cell(row = count, column = 2).value)
            obj.setIDPERCEPCIONES(sheet_percepcion.cell(row = count, column = 10).value)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,(sheet_percepcion.cell(row = count, column = 1).value),True))
            obj.setTIPOPERCEPCION(sheet_percepcion.cell(row = count, column = 3).value)
            obj.setCLAVE(sheet_percepcion.cell(row = count, column = 4).value)
            obj.setCONCEPTO(sheet_percepcion.cell(row = count, column = 5).value)
            obj.setIMPORTEGRAVADO(sheet_percepcion.cell(row = count, column = 6).value)
            obj.setIMPORTEEXENTO(sheet_percepcion.cell(row = count, column = 7).value)
            obj.setIMPORTETOTAL(sheet_percepcion.cell(row = count, column = 8).value)
            obj.setFEC_IMPUTACION(sheet_percepcion.cell(row = count, column = 9).value)
            obj.setFEC_PAGO(sheet_complemento.cell(row = count, column = 3).value)
            obj.setFEC_INICIO_P(sheet_complemento.cell(row = count, column = 4).value)
            obj.setFEC_FIN_P(sheet_complemento.cell(row = count, column = 5).value)
            Listpercepcion.append(obj)
            idPercepcion = idPercepcion + 1
            count = count + 1
            
        dc = open(fileaux,"w")
        
        cad = "IDPERCEPCION,IDPERCEPCIONES,IDNOMINA,TIPOPERCEPCION,CLAVE,CONCEPTO,IMPORTEGRAVADO,IMPORTEEXENTO,IMPORTETOTAL,FEC_IMPUTACION,FEC_PAGO,FEC_INICIO_P,FEC_FIN_P,IDUSUARIO,FG,ST,SERIE\n"
        dc.write(cad) 
        for lista in Listpercepcion:
            dc.write(lista.getinfo(1))
        dc.close()
        #----------------------------------------FIN DE PERCEPCION -------------------------------------------    
        
    #----------------------------------------INICIO DE DEDUCCIONES -------------------------------------------
    fileaux = path + "/M4T_DEDUCCIONES_CND_33.csv"
    count =  7
    Listdeducciones = []
    idDeducciones = 1
        
    if sheet_deduccion.cell(row = count, column = 1).value != None:
        while sheet_deduccion.cell(row = count, column = 1).value != None:
            obj = Deducciones()
            obj.setIDDEDUCCIONES(idDeducciones)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_deduccion.cell(row = count, column = 1).value,True))
            obj.setTOTALOTRASDEDUCCIONES(sheet_deduccion.cell(row = count, column = 2).value)
            obj.setTOTALIMPUESTOSRETENIDOS(sheet_deduccion.cell(row = count, column = 3).value)
            
            Listdeducciones.append(obj)
            idDeducciones = idDeducciones + 1
            count = count + 1
            
        dc = open(fileaux,"w")
        
        cad = "IDDEDUCCIONES,IDNOMINA,TOTALOTRASDEDUCCIONES,TOTALIMPUESTOSRETENIDOS,IDUSUARIO,FG,ST\n"
        dc.write(cad) 
        for lista in Listdeducciones:
            dc.write(lista.getinfo(1))
        dc.close()
        #----------------------------------------FIN DE DEDUCCIONES -------------------------------------------
    #----------------------------------------INICIO DE DEDUCCION -------------------------------------------
    fileaux = path + "/M4T_DEDUCCION_NDD_33.csv"
    count =  7
    Listdeduccion = []
    idDeduccion = 1
    
    if sheet_deducciones.cell(row = count, column = 1).value != None:
        while sheet_deducciones.cell(row = count, column = 1).value != None:
            obj = Deduccion()
            obj.setIDDEDUCCION(sheet_deducciones.cell(row = count, column = 7).value)
            obj.setIDDEDUCCIONES(sheet_deducciones.cell(row = count, column = 8).value)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_deducciones.cell(row = count, column = 1).value,True))
            obj.setTIPODEDUCCION(sheet_deducciones.cell(row = count, column = 2).value)
            obj.setCLAVE(sheet_deducciones.cell(row = count, column = 3).value)
            obj.setCONCEPTO(sheet_deducciones.cell(row = count, column = 4).value)
            obj.setIMPORTE(sheet_deducciones.cell(row = count, column = 5).value)
            obj.setFEC_IMPUTACION(sheet_deducciones.cell(row = count, column = 6).value)
        
            Listdeduccion.append(obj)
            idDeduccion = idDeduccion + 1
            count = count + 1
        
        dc = open(fileaux,"w")
        
        cad = "IDDEDUCCION,IDDEDUCCIONES,IDNOMINA,TIPODEDUCCION,CLAVE,CONCEPTO,IMPORTE,FEC_IMPUTACION,FEC_PAGO,TIPO_PRESTAMO,SUBTIPO_PRESTAMO,IDUSUARIO,FG,ST,SERIE,IMPORTEGRAVADO,IMPORTEEXENTO\n"
        dc.write(cad) 
        for lista in Listdeduccion:
            dc.write(lista.getinfo(1))
        dc.close()
        #----------------------------------------FIN DE DEDUCCION -------------------------------------------        
    #----------------------------------------INICIO DE OtrosPagos -------------------------------------------        
    fileaux = path + "/M4T_OTROSPAGOS_NOP_33.csv"
    count =  7
    ListOtrosPagos = []
    idOtrosPagos = 1
        
    if sheet_pagos.cell(row = count, column = 1).value != None:
        while sheet_pagos.cell(row = count, column = 1).value != None:
            obj = OtrosPagos()
            obj.setIDOTROPAGO(idOtrosPagos)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_pagos.cell(row = count, column = 1).value,True))
            obj.setTIPOOTROPAGO(sheet_pagos.cell(row = count, column = 2).value)
            obj.setCLAVE(sheet_pagos.cell(row = count, column = 3).value)
            obj.setCONCEPTO(sheet_pagos.cell(row = count, column = 4).value)
            obj.setIMPORTE(sheet_pagos.cell(row = count, column = 5).value)
            obj.setSUBSIDIOCAUSADO(sheet_pagos.cell(row = count, column = 6).value)
            obj.setSALDOAFAVOR(sheet_pagos.cell(row = count, column = 7).value)
            obj.setANIO(sheet_pagos.cell(row = count, column = 8).value)
            obj.setREMANENTESALDOAFAVOR(sheet_pagos.cell(row = count, column = 9).value)
            
            ListOtrosPagos.append(obj)
            idOtrosPagos = idOtrosPagos + 1
            count = count + 1
            
        dc = open(fileaux,"w")
        
        cad = "IDOTROPAGO,IDNOMINA,TIPOOTROPAGO,CLAVE,CONCEPTO,IMPORTE,SUBSIDIOCAUSADO,SALDOAFAVOR,ANIO,REMANENTESALDOAFAVOR,IDUSUARIO,FG,ST\n"
        dc.write(cad) 
        for lista in ListOtrosPagos:
            dc.write(lista.getinfo(1))
        dc.close()
        #----------------------------------------FIN DE OtrosPagos ------------------------------------------- 
    #----------------------------------------INICIO DE Incapacidades ------------------------------------------- 
    fileaux = path + "/M4T_INCAPACIDADES_CNI_33.csv"
    count =  7
    ListIncapacidades = []
    idIncapacidades = 1
        
    if sheet_incapacidad.cell(row = count, column = 1).value != None:
        while sheet_incapacidad.cell(row = count, column = 1).value != None:
            obj = Incapacidades()
            obj.setIDINCAPACIDAD(idIncapacidades)
            obj.setIDNOMINA(buscarEmpleado(listEmpleados,sheet_incapacidad.cell(row = count, column = 1).value,True))
            obj.setDIASINCAPACIDAD(sheet_incapacidad.cell(row = count, column = 2).value)
            obj.setTIPOINCAPACIDAD(sheet_incapacidad.cell(row = count, column = 3).value)
            obj.setIMPORTEMONETARIO(sheet_incapacidad.cell(row = count, column = 4).value)
                            
            
            ListIncapacidades.append(obj)
            idIncapacidades = idIncapacidades + 1
            count = count + 1
            
        dc = open(fileaux,"w")
        
        cad = "IDINCAPACIDAD,IDNOMINA,DIASINCAPACIDAD,TIPOINCAPACIDAD,IMPORTEMONETARIO,IDUSUARIO,FG,ST\n"
        dc.write(cad) 
        for lista in ListIncapacidades:
            dc.write(lista.getinfo(1))
        dc.close()
        #---------------------------------------- FIN DE Incapacidades ------------------------------------------- 
    
    
    #Cierre de archivos
    wb.close()

espacio = '-'
now = datetime.now()

print("Digite el nombre del archivo")

file = "fileprueba/" + input() + ".xlsx"
#file = "fileprueba/RTP0001071K1-I1010A10PDRT-B42_20-00-.xlsx" 
cut = file.split('-', 2)
folder = "out/" + str(format(now.year) + espacio + format(now.month) + espacio + format(now.day) + '_' + cut[2])
nombrezip = folder[:-6]
#print(nombrezip)

def crear_zip(folder):
    generar_zip = zipfile.ZipFile(nombrezip + '.zip', 'w')

    for folder, subfolders, files in os.walk(folder):

        for file in files:
            if file.endswith('.csv'):
                generar_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder,file), folder), compress_type = zipfile.ZIP_DEFLATED)

    generar_zip.close()
    print('Proceso completado')

if not os.path.isdir(folder):
    os.mkdir(folder)
    crearArchivos(folder, file)
    crear_zip(folder)
else: 
    while True:
        print("¿Carpeta ya existe, deseas continuar?(y/n)")
        resp = input()
        if resp == "y":
            crearArchivos(folder, file)
            crear_zip(folder)
            break
        if resp == "n":
            break
        if resp != "y" and resp != "n":
            print("Opcion invalida")

 